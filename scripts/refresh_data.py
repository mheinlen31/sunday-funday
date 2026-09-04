#!/usr/bin/env python3
"""Refresh Sunday Funday keeper data.

Reads rosters from 'Sunday Funday Stuff.xlsx' (2026 Keeper Values sheet),
fetches current ESPN Live Draft Results average auction values, recomputes
keeper prices per the Manifesto, and writes js/data.js for the site.

Pricing rules (Manifesto v5.1):
  - FA pickup ('-' draft cost): keeper price = current market value
  - Drafted/kept player: market up more than $10 -> prev cost + $10
                         otherwise -> (prev cost + market) / 2
  - Repeat keepers (Acquired Via "Keeper...") kept again sign 2-yr deals:
    next year = this year's price + $5
  - Rows whose price cell is hardcoded in the sheet are contract-locked
    (renewals at prev + $5, or year 2 marked CONTRACT); ESPN values ignored.

Run:  python3 scripts/refresh_data.py
"""
import json
import math
import re
import unicodedata
import urllib.request
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "Sunday Funday Stuff.xlsx"
SHEET = "2026 Keeper Values"
SEASON = 2026
BAND_ROWS = (1, 21, 41, 61, 81)  # header row of each two-team band
BLOCK_COLS = (1, 9)              # first column of the left/right team block

# Keeper values lock at Sept 2, 2026, noon Central (CDT = UTC-5) -> 17:00 UTC.
# After this, values are FINAL: don't query ESPN, reuse the frozen values from
# the last pre-lock data.js. Trades (roster changes) still apply — a player
# carries his frozen value to his new team.
VALUES_LOCK = datetime(2026, 9, 2, 17, 0, tzinfo=timezone.utc)

ESPN_URL = (
    "https://lm-api-reads.fantasy.espn.com/apis/v3/games/ffl/seasons/"
    f"{SEASON}/segments/0/leaguedefaults/3?view=kona_player_info"
)
ESPN_FILTER = {"players": {"limit": 2500, "sortDraftRanks": {
    "sortPriority": 100, "sortAsc": True, "value": "STANDARD"}}}
POS_IDS = {1: "QB", 2: "RB", 3: "WR", 4: "TE", 5: "K", 16: "D/ST"}
SUFFIXES = {"jr", "sr", "ii", "iii", "iv", "v"}
# sheet spelling -> ESPN spelling
ALIASES = {
    "rickey pearsall": "ricky pearsall",
    "kenneth gainwell": "kenny gainwell",
    "jason meyers": "jason myers",
}
# Budgets-sheet team name -> Keeper Values sheet team name
TEAM_ALIASES = {"Chovies": "The Chovies", "HopefulChovies": "The Chovies",
                "AFRESHAYPEPPER ASAYWHEN": "Pep", "Gorlock": "Loser"}
# 2026 base auction purses confirmed by Matt 2026-07-10 (pre-trade); win over
# the Budgets sheet. Draft-dollar trades adjust these via TRADES below.
BASE_PURSES = {
    "The Chovies": 197, "Pep": 201, "Bom Bers": 199, "Centersup": 201,
    "Chance": 202, "Juice": 200, "Loser": 196, "Magic Rats": 208,
    "Paw": 196, "Silent Pugios": 200,
}
START_PURSE = 200      # every team's 2026 auction budget before any trade

# Draft dollars from earlier seasons that land on the 2026 purse, transcribed
# from the workbook's trade history. "pays" loses the dollars, "gets" receives
# them; "player" is the asset that moved from gets -> pays. Checked against
# BASE_PURSES at import time (below), so the itemised ledger and the confirmed
# purses can never silently disagree.
#
# "The Fighting Mamdanis" is the 2025 name of the franchise that also went by
# Gorlock the Destroyer and now by Loser -- Josh Allen sits on Gorlock's
# end-of-2025 roster, which is the 08/15 trade here. Those two Mamdanis lines
# net to zero for that team, so every team's TOTAL is right either way; only
# the itemisation depends on the identification.
PRIOR_2026_MOVES = [
    {"date": "2024-12-02", "amount": 1, "pays": "Paw", "gets": "Loser",
     "player": "Baker Mayfield"},
    {"date": "2025-08-13", "amount": 2, "pays": "Paw", "gets": "Bom Bers",
     "player": "Ja'Marr Chase"},
    {"date": "2025-08-15", "amount": 1, "pays": "Magic Rats", "gets": "Pep",
     "player": "James Cook"},
    {"date": "2025-08-15", "amount": 1, "pays": "Loser", "gets": "Bom Bers",
     "player": "Josh Allen"},
    {"date": "2025-08-16", "amount": 1, "pays": "The Chovies", "gets": "Centersup",
     "player": "Saquon Barkley", "extra": "James Conner went back the other way"},
    {"date": "2025-08-19", "amount": 1, "pays": "Paw", "gets": "Loser",
     "player": "Travis Etienne"},
    {"date": "2025-08-22", "amount": 2, "pays": "The Chovies", "gets": "Chance",
     "player": "Ricky Pearsall"},
    {"date": "2025-11-24", "amount": 4, "pays": "Bom Bers", "gets": "Loser",
     "player": "Wan'Dale Robinson"},
    {"date": "2025-12-16", "amount": 9, "pays": "Loser", "gets": "Magic Rats",
     "player": "James Cook"},
]


def _check_prior_moves():
    """$200 + the prior-season moves must reproduce every confirmed purse."""
    tally = {n: float(START_PURSE) for n in BASE_PURSES}
    for m in PRIOR_2026_MOVES:
        for side in ("pays", "gets"):
            if m[side] not in tally:
                raise SystemExit(f"\nBUDGET ERROR: unknown team {m[side]!r} in "
                                 f"PRIOR_2026_MOVES ({m['date']}).\n")
        tally[m["pays"]] -= m["amount"]
        tally[m["gets"]] += m["amount"]
    bad = {n: (tally[n], BASE_PURSES[n]) for n in BASE_PURSES
           if tally[n] != BASE_PURSES[n]}
    if bad:
        lines = "\n".join(f"  {n}: ledger ${a:g} vs confirmed purse ${b}"
                           for n, (a, b) in sorted(bad.items()))
        raise SystemExit(
            "\nBUDGET ERROR: PRIOR_2026_MOVES no longer reconciles with "
            f"BASE_PURSES.\n{lines}\n  Nothing was written. Fix one or the "
            "other and re-run.\n")


_check_prior_moves()


# Offseason 2026 trades. Each side's "players"/"dollars" is what it GIVES UP;
# in a two-team trade those go to the other side. Roster moves are applied to
# the rosters read from the workbook (a keeper's status/cost follows him) and
# purse changes are derived from the dollars — the workbook is left untouched so
# its historical formula caches stay intact. "grade"/"note" are my hand-authored
# take per side; the site also shows the objective value each side received.
TRADES = [
    {
        "date": "2026-08-16",
        "summary": "Paw traded Kenneth Gainwell to Centersup for $2 in 2026 draft budget.",
        "sides": {
            "Paw": {"players": ["Kenneth Gainwell"], "dollars": 0, "grade": "B+",
                    "note": "A $3 body who was never making a $100 keeper board — Chase and "
                            "St. Brown lock $95 of Paw's cap before they choose anything. "
                            "Gainwell was worth nothing to this roster, so $2 is free, if "
                            "only barely worth the paperwork."},
            "Centersup": {"players": [], "dollars": 2, "grade": "B-",
                    "note": "Two dollars for depth that doesn't crack their keeper set. "
                            "Gainwell adds nothing to the best $100 build Centersup can "
                            "make, so this is money spent on a player headed for the "
                            "auction pool anyway. Harmless at the price."},
        },
    },
    {
        "date": "2026-08-25",
        "summary": "Paw traded Christian Watson to the Chovies for $3 in 2026 and $1 in 2027.",
        "sides": {
            "Paw": {"players": ["Christian Watson"], "dollars": 0, "dollars2027": 0,
                    "grade": "A-",
                    "note": "Another asset Paw could never use. With $95 of the cap "
                            "mandatory in Chase and St. Brown, Watson's $6 keeper price was "
                            "surplus they had no room to bank — his marginal worth to this "
                            "roster is zero. Four dollars across two drafts for nothing is "
                            "clean profit."},
            "The Chovies": {"players": [], "dollars": 3, "dollars2027": 1, "grade": "C+",
                    "note": "Paid $4 for about $0.40 of keeper value. Watson joins a "
                            "receiver room already deep in interchangeable $1-6 fliers — "
                            "Hunter, Pearsall, Boutte, Dike, Horton — and doesn't improve "
                            "the best set the Chovies can build. Cheap, but money for "
                            "nothing."},
        },
    },
    {
        "date": "2026-08-17",
        "summary": "Loser traded the rights to Ashton Jeanty to the Magic Rats for $3 in 2026 and $1 in 2027.",
        "sides": {
            "Loser": {"players": ["Ashton Jeanty"], "dollars": 0, "dollars2027": 0,
                    "grade": "A-",
                    "note": "The best kind of sale — an asset worth zero to the seller. "
                            "Loser's optimal board is Chase Brown ($17), Cook ($29), "
                            "Judkins, McConkey, Waddle and Maye: six keepers, $99, roughly "
                            "+$70 of surplus. Jeanty at $45 never came close to making it. "
                            "They turned a player they were always releasing into $4."},
            "Magic Rats": {"players": [], "dollars": 3, "dollars2027": 1, "grade": "C-",
                    "note": "The purchase that constrained everything after it. Jeanty is "
                            "priced $7 above his own ADP, so he carries negative surplus — "
                            "and because he came by trade he is effectively committed at "
                            "$45, which is precisely what later made the London buy "
                            "redundant. Paying $4 for the right to spend $45 on a "
                            "below-market keeper is a bet on the player alone, with the cap "
                            "math running against it."},
        },
    },
    {
        "date": "2026-09-02",
        "summary": "The Chovies traded Dylan Sampson, $1 in 2026 and $1 in 2027 "
                   "to Chance for Quentin Johnston.",
        "sides": {
            "The Chovies": {"players": ["Dylan Sampson"], "dollars": 1, "dollars2027": 1,
                    "grade": "C+",
                    "note": "Two dollars for a lateral. Both men are $1-2 keepers sitting "
                            "exactly at market and neither touches the Chovies' best "
                            "build, so the cash is the only thing that genuinely moves. "
                            "Paying for a preference."},
            "Chance": {"players": ["Quentin Johnston"], "dollars": 0, "dollars2027": 0,
                    "grade": "B+",
                    "note": "Took the money on a swap of equivalents. Neither player makes "
                            "any keeper board, so being paid $2 to exchange one fringe body "
                            "for another is the right side of a nothing trade — and Sampson "
                            "lands in a backfield whose two worst contracts are Kamara ($11 "
                            "for a $1 player) and Aaron Jones."},
        },
    },
    {
        "date": "2026-09-02",
        "summary": "Pep traded Omarion Hampton to Juice for $4 in 2026.",
        "sides": {
            "Pep": {"players": ["Omarion Hampton"], "dollars": 0, "grade": "B+",
                    "note": "Sold into a forced choice and got paid for it. Achane's $40 is "
                            "a mandatory year-two, and Hampton ($39) and Walker ($34) were "
                            "never both affordable underneath it — Hampton was worth only "
                            "about $1 more to Pep than the next-best use of that money. "
                            "Converting him into $4 of auction cash is a clear gain."},
            "Juice": {"players": [], "dollars": 4, "grade": "C+",
                    "note": "A fair player at a slightly rich price. Hampton does make "
                            "their best set — Hampton, Javonte, McMillan and Bowers for "
                            "$92 — but only about $1.50 better than the alternative build, "
                            "so $4 is a couple of dollars of overpay. Not damaging, just "
                            "not the bargain that +$4 on ADP implies."},
        },
    },
    {
        "date": "2026-09-04",
        "summary": "Paw traded Travis Etienne Jr. to Centersup for $2 in 2026 "
                   "and $2 in 2027.",
        "sides": {
            "Paw": {"players": ["Travis Etienne Jr."], "dollars": 0, "dollars2027": 0,
                    "grade": "A-",
                    "note": "The cleanest sale of the offseason. Chase and St. Brown lock "
                            "$95 of Paw's $100 before any choice, so Etienne's +$9 was "
                            "unusable — keeping him meant $7 over the cap and $14 of "
                            "luxury tax. He was worth literally nothing to this roster and "
                            "returned $4. The only ding is that a capped-out seller has no "
                            "leverage, and the league knew it."},
            "Centersup": {"players": [], "dollars": 2, "dollars2027": 2, "grade": "A-",
                    "note": "The buyer with the room the seller lacked. Etienne is worth "
                            "about +$9 to Centersup's best build — Gibbs, Jefferson, "
                            "Pickens, Etienne, Warren and Gainwell — for a $4 outlay. That "
                            "is the largest genuine value gain in any trade this offseason, "
                            "and he fills the one real hole: behind Gibbs they were running "
                            "Hunt, Vidal and Croskey-Merritt."},
        },
    },
    {
        "date": "2026-09-04",
        "summary": "Silent Pugios traded Drake London to the Magic Rats for $7 in "
                   "2026 and $2 in 2027.",
        "sides": {
            "Silent Pugios": {"players": ["Drake London"], "dollars": 0, "dollars2027": 0,
                    "grade": "B+",
                    "note": "Sold at par and pocketed the structure. Their best set with "
                            "London — Smith-Njigba, Swift, Flowers, Loveland, London — is "
                            "worth about +$66 in real auction dollars; without him, "
                            "Smith-Njigba, Taylor, Flowers and Harrison is about +$57. So "
                            "London was worth roughly $8 here and they were paid $9. Even "
                            "on value, and they kept 2027 clear of a second locked receiver "
                            "(London renews at $35 on a roster already carrying "
                            "Smith-Njigba). Taylor at $58 also reads better than +$1, since "
                            "a $59 ADP player costs about $63 once keeper surplus inflates "
                            "the auction."},
            "Magic Rats": {"players": [], "dollars": 7, "dollars2027": 2, "grade": "C",
                    "note": "Nine dollars for nothing this year. With Jeanty committed at "
                            "$45, Jeanty plus London plus Lamb is $116 and $32 of tax — "
                            "London can't join Lamb, only replace him, and the two builds "
                            "are worth the same. The entire return is 2027: London renews "
                            "at $35 where Lamb renews at $46, and he's three years younger. "
                            "A real plan, but an expensive way to reach it, and it likely "
                            "sends a $51 ADP receiver to the auction."},
        },
    },
]
# NFL team name -> ESPN logo abbreviation (for D/ST images)
NFL_ABBR = {
    "cardinals": "ari", "falcons": "atl", "ravens": "bal", "bills": "buf",
    "panthers": "car", "bears": "chi", "bengals": "cin", "browns": "cle",
    "cowboys": "dal", "broncos": "den", "lions": "det", "packers": "gb",
    "texans": "hou", "colts": "ind", "jaguars": "jax", "chiefs": "kc",
    "raiders": "lv", "chargers": "lac", "rams": "lar", "dolphins": "mia",
    "vikings": "min", "patriots": "ne", "saints": "no", "giants": "nyg",
    "jets": "nyj", "eagles": "phi", "steelers": "pit", "49ers": "sf",
    "seahawks": "sea", "buccaneers": "tb", "titans": "ten", "commanders": "wsh",
}
HEADSHOT = "https://a.espncdn.com/i/headshots/nfl/players/full/{}.png"
TEAM_LOGO = "https://a.espncdn.com/i/teamlogos/nfl/500/{}.png"
# ESPN proTeamId -> NFL abbreviation (0 = free agent / none)
PRO_TEAM = {
    1: "ATL", 2: "BUF", 3: "CHI", 4: "CIN", 5: "CLE", 6: "DAL", 7: "DEN",
    8: "DET", 9: "GB", 10: "TEN", 11: "IND", 12: "KC", 13: "LV", 14: "LAR",
    15: "MIA", 16: "MIN", 17: "NE", 18: "NO", 19: "NYG", 20: "NYJ", 21: "PHI",
    22: "ARI", 23: "PIT", 24: "LAC", 25: "SF", 26: "SEA", 27: "TB", 28: "WSH",
    29: "CAR", 30: "JAX", 33: "BAL", 34: "HOU",
}


def norm_name(name):
    s = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-z ]", "", s.lower().replace(".", " ").replace("'", ""))
    parts = [p for p in s.split() if p not in SUFFIXES]
    return " ".join(parts)


def fetch_espn():
    req = urllib.request.Request(ESPN_URL, headers={
        "X-Fantasy-Filter": json.dumps(ESPN_FILTER),
        "User-Agent": "Mozilla/5.0",
    })
    with urllib.request.urlopen(req, timeout=60) as r:
        data = json.load(r)
    market = {}
    for entry in data.get("players", []):
        p = entry.get("player") or {}
        own = p.get("ownership") or {}
        aav = own.get("auctionValueAverage")
        pos = POS_IDS.get(p.get("defaultPositionId"))
        if aav is None or pos is None:
            continue
        name = p.get("fullName", "")
        if pos == "D/ST":
            name = name.replace(" D/ST", "")
        key = (norm_name(name), pos)
        # keep the highest AAV if a name collides
        if key not in market or aav > market[key]["aav"]:
            market[key] = {"aav": aav, "id": p.get("id"), "name": name,
                           "nfl": PRO_TEAM.get(p.get("proTeamId"))}
    return market


def build_pool(market, owned_ids, owned_names, limit=300):
    """Unowned ESPN players worth showing on the ADP board.

    Ownership is checked by name as well as ESPN id: a frozen market rebuilt
    from data.js has no id for D/ST (team-logo URLs carry none), so an
    id-only check would put None in owned_ids and hide every free defense.
    """
    pool = []
    for (key, pos), e in market.items():
        if key in owned_names or e["aav"] < 0.5:
            continue
        if e["id"] is not None and e["id"] in owned_ids:
            continue
        pool.append({"name": e["name"], "pos": pos,
                     "market": adp_round(e["aav"]),
                     "nfl": e.get("nfl") if pos != "D/ST" else None,
                     "img": player_img(e["name"], pos, e["id"])})
    pool.sort(key=lambda p: -p["market"])
    return pool[:limit]


def lookup(market, name, pos):
    """Return (entry, espn_pos) matching by name+pos, then alias, then name only."""
    n = norm_name(name)
    n = ALIASES.get(n, n)
    if (n, pos) in market:
        return market[(n, pos)], pos
    for (mn, mp), entry in market.items():
        if mn == n:
            return entry, mp
    return None, pos


def player_img(name, pos, espn_id):
    if pos == "D/ST":
        abbr = NFL_ABBR.get(norm_name(name))
        return TEAM_LOGO.format(abbr) if abbr else None
    return HEADSHOT.format(espn_id) if espn_id else None


def fmt_money(x):
    return int(x) if x == int(x) else x


def adp_round(raw):
    """ESPN's raw auctionValueAverage -> the whole-dollar ADP the league uses.

    Two steps, and the order matters. ESPN's site shows one decimal, and the
    league has always rounded the number on the screen -- so round to ESPN's
    displayed precision FIRST, then apply standard half-up rounding (0-4 down,
    5-9 up). Bijan at a raw 68.47 shows as 68.5 on ESPN and is a $69 ADP to
    us; rounding the full-precision value straight to a dollar would say $68.

    Note this is NOT the round-down rule -- that one applies later, to the
    keeper-price average, which lands on .0 or .5 precisely because this
    function hands it a whole dollar.

    Python's built-in round() is no good here twice over: it rounds half to
    EVEN (round(16.5) == 16, round(17.5) == 18), and it would be working off
    full precision rather than what ESPN displays.
    """
    return max(1, int(math.floor(round(raw, 1) + 0.5)))


def keeper_price(draft_cost, market_val):
    if draft_cost is None:  # FA pickup
        return market_val
    if market_val - draft_cost > 10:
        return draft_cost + 10
    return (draft_cost + market_val) / 2


def read_rosters():
    wb_vals = load_workbook(XLSX, data_only=True)
    wb_forms = load_workbook(XLSX)
    wsv, wsf = wb_vals[SHEET], wb_forms[SHEET]
    teams = []
    for band in BAND_ROWS:
        for col in BLOCK_COLS:
            name = wsv.cell(band, col).value
            if not name:
                continue
            players = []
            for r in range(band + 1, band + 20):
                pname = wsv.cell(r, col).value
                if pname is None:
                    break
                raw_cost = wsv.cell(r, col + 3).value
                price_cell = wsf.cell(r, col + 5).value
                players.append({
                    "name": str(pname).strip(),
                    "pos": wsv.cell(r, col + 1).value,
                    "acquired": wsv.cell(r, col + 2).value,
                    "draftCost": None if raw_cost in ("-", None) else float(raw_cost),
                    "sheetMarket": wsv.cell(r, col + 4).value,
                    "sheetPrice": wsv.cell(r, col + 5).value,
                    "priceLocked": not (isinstance(price_cell, str) and price_cell.startswith("=")),
                    "contractCell": wsv.cell(r, col + 6).value,
                })
            teams.append({"name": str(name).strip(), "players": players})
    return teams


def frozen_market(prev):
    """A lookup()-compatible market map rebuilt from a locked data.js, so
    post-lock trades keep the final frozen values instead of new ESPN ADP."""
    m = {}

    def add(name, pos, market_val, img, nfl):
        eid = None
        if img:
            mo = re.search(r"/full/(\d+)\.png", img)
            if mo:
                eid = int(mo.group(1))
        # key the same way lookup() searches (alias-normalized) so the three
        # sheet/ESPN spelling mismatches still resolve
        key = ALIASES.get(norm_name(name), norm_name(name))
        m[(key, pos)] = {"aav": market_val, "id": eid, "name": name, "nfl": nfl}

    for t in (prev or {}).get("teams", []):
        for p in t["players"]:
            add(p["name"], p["pos"], p["market"], p.get("img"), p.get("nfl"))
    for p in (prev or {}).get("pool", []):
        add(p["name"], p["pos"], p["market"], p.get("img"), p.get("nfl"))
    return m


HISTORY_YEARS = range(2025, 2018, -1)  # 2025 down to 2019


def read_history():
    """Past-season rosters from the older '<year> Keeper Values' sheets.

    Block layout (header row, name column) varies year to year, so team
    blocks are found dynamically by locating 'Position' header cells. Columns
    relative to the name column are stable: +1 pos, +2 acquired, +3 prior
    draft cost, +5 keeper price. Row-major block order matches the current
    roster order, so a block's index is its stable franchise slot.
    """
    wb = load_workbook(XLSX, data_only=True)

    def num(v):
        try:
            return fmt_money(float(v))
        except (TypeError, ValueError):
            return None

    by_year = {}
    for year in HISTORY_YEARS:
        name = f"{year} Keeper Values"
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        headers = []  # (header_row, name_col), row-major
        for row in ws.iter_rows(min_row=1, max_row=160):
            for c in row:
                if c.value == "Position" and c.column > 1:
                    headers.append((c.row, c.column - 1))
        teams = []
        for slot, (hr, nc) in enumerate(headers):
            tname = ws.cell(hr, nc).value
            if not tname:
                continue
            players = []
            for r in range(hr + 1, hr + 25):
                pn = ws.cell(r, nc).value
                if pn is None:
                    break
                raw_cost = ws.cell(r, nc + 3).value
                players.append({
                    "name": str(pn).strip(),
                    "pos": ws.cell(r, nc + 1).value,
                    "acquired": ws.cell(r, nc + 2).value,
                    "draftCost": None if raw_cost in ("-", None) else num(raw_cost),
                    "price": num(ws.cell(r, nc + 5).value),
                })
            teams.append({"name": str(tname).strip(), "slot": slot, "players": players})
        by_year[str(year)] = teams
    return {"seasons": [y for y in HISTORY_YEARS if str(y) in by_year],
            "byYear": by_year}


def load_previous():
    """Previous data.js payload, for diffing prices into the change log."""
    p = ROOT / "js" / "data.js"
    if not p.exists():
        return None
    s = p.read_text()
    try:
        return json.loads(s[s.index("{"):s.rindex("}") + 1])
    except (ValueError, json.JSONDecodeError):
        return None


def update_changelog(prev, teams):
    """Append today's keeper-price changes; keep the whole offseason
    (capped at the most recent 500 entries)."""
    log = list((prev or {}).get("changeLog", []))
    today = datetime.now(timezone.utc).date()
    prev_prices = {(t["name"], p["name"]): p["price"]
                   for t in (prev or {}).get("teams", []) for p in t["players"]}
    for t in teams:
        for p in t["players"]:
            old = prev_prices.get((t["name"], p["name"]))
            if old is None or old == p["price"]:
                continue
            same_day = next((e for e in log if e["d"] == today.isoformat()
                             and e["team"] == t["name"] and e["name"] == p["name"]), None)
            if same_day:  # second refresh today: keep original 'from'
                same_day["to"] = p["price"]
                if same_day["from"] == same_day["to"]:
                    log.remove(same_day)
            else:
                log.append({"d": today.isoformat(), "team": t["name"],
                            "name": p["name"], "pos": p["pos"],
                            "from": old, "to": p["price"]})
    return log[-500:]


def read_purses():
    """Auction purse per team: confirmed base purses, then draft-dollar trades."""
    purses = {name: float(v) for name, v in BASE_PURSES.items()}
    for tr in TRADES:
        names = list(tr["sides"])
        for i, name in enumerate(names):
            other = names[1 - i] if len(names) == 2 else None
            d = tr["sides"][name].get("dollars", 0)
            if d and other:
                purses[name] -= d          # gives dollars away
                purses[other] += d          # other side receives them
    return purses


def build_budgets(purses):
    """Per-team ledgers for both draft years, $200 -> today's budget.

    2026 rows come from PRIOR_2026_MOVES plus this offseason's cash; 2027 rows
    are the forward dollars traded so far. Rows carry a running balance so the
    page can show the arithmetic rather than just the endpoints, and both years
    fall out of TRADES, so a new trade appears in the right one for free.
    """
    rows = {SEASON: {n: [] for n in BASE_PURSES},
            SEASON + 1: {n: [] for n in BASE_PURSES}}

    def add(year, team, other, date, delta, why, note=None, season=None):
        rows[year][team].append({"date": date, "delta": fmt_money(delta),
                                 "with": other, "why": why, "note": note,
                                 "season": season})

    for m in PRIOR_2026_MOVES:
        yr = int(m["date"][:4])
        add(SEASON, m["pays"], m["gets"], m["date"], -m["amount"],
            f'Paid for {m["player"]}', m.get("extra"), yr)
        add(SEASON, m["gets"], m["pays"], m["date"], m["amount"],
            f'Sold {m["player"]}', m.get("extra"), yr)

    for tr in TRADES:
        names = list(tr["sides"])
        if len(names) != 2:
            continue
        for i, name in enumerate(names):
            other = names[1 - i]
            # what this side received is what the other side gave up
            got = tr["sides"][other].get("players") or []
            label = ", ".join(got)
            for year, key in ((SEASON, "dollars"), (SEASON + 1, "dollars2027")):
                d = tr["sides"][name].get(key, 0)
                if not d:
                    continue
                add(year, name, other, tr["date"], -d,
                    f"Paid for {label}" if label else f"Cash to {other}", None, SEASON)
                add(year, other, name, tr["date"], d,
                    f"Sold {label}" if label else f"Cash from {name}", None, SEASON)

    out = []
    for name in BASE_PURSES:
        years = {}
        for year in (SEASON, SEASON + 1):
            r = sorted(rows[year][name], key=lambda x: (x["date"], -abs(x["delta"])))
            bal = float(START_PURSE)
            for x in r:
                bal += x["delta"]
                x["balance"] = fmt_money(bal)
            years[str(year)] = {"rows": r, "final": fmt_money(bal),
                                "net": fmt_money(bal - START_PURSE)}
        # the confirmed purse is the authority for this season; the ledger must
        # land on it or the itemisation is lying about where the money went
        if years[str(SEASON)]["final"] != fmt_money(purses[name]):
            raise SystemExit(
                f"\nBUDGET ERROR: {name} {SEASON} ledger ends at "
                f'${years[str(SEASON)]["final"]} but the purse is ${purses[name]:g}. '
                "Nothing was written.\n")
        out.append({"team": name, "start": START_PURSE, "years": years,
                    # kept flat for the current season so existing views work
                    "rows": years[str(SEASON)]["rows"],
                    "final": years[str(SEASON)]["final"],
                    "net": years[str(SEASON)]["net"],
                    "next": years[str(SEASON + 1)]["net"]})
    return out


def apply_trades(teams):
    """Move traded players to the other side (keeper status/cost follows them)."""
    by_name = {t["name"]: t for t in teams}
    for tr in TRADES:
        names = list(tr["sides"])
        if len(names) != 2:
            print(f"trade warning: {tr.get('summary')} — not a two-team trade, skipped")
            continue
        for i, name in enumerate(names):
            src, dst = by_name.get(name), by_name.get(names[1 - i])
            if not src or not dst:
                print(f"trade warning: team not found ({name})")
                continue
            for player in tr["sides"][name].get("players", []):
                moving = [p for p in src["players"]
                          if norm_name(p["name"]) == norm_name(player)]
                if not moving:
                    # Half-applying is worse than failing: the dollars would
                    # still move (read_purses works off the same TRADES list)
                    # and the league would silently go out of balance.
                    raise SystemExit(
                        f"\nTRADE ERROR: '{player}' is not on {name}.\n"
                        f"  Trade: {tr.get('summary')}\n"
                        f"  Nothing was written. Fix the team name or player "
                        f"spelling in TRADES and re-run.\n")
                for p in moving:
                    src["players"].remove(p)
                    dst["players"].append(p)
    for tr in TRADES:
        print(f"trade applied: {tr['summary']}")


def read_trade_history():
    """Narrative trade log from the workbook, grouped by year (newest first)."""
    ws = load_workbook(XLSX, data_only=True)["Sunday Funday Trade History"]
    groups, cur = [], None
    for row in ws.iter_rows(max_col=1):
        v = row[0].value
        if v is None:
            continue
        s = str(v).strip()
        m = re.match(r"^(\d{4})\s+Trades$", s)
        if m:
            cur = {"year": int(m.group(1)), "entries": []}
            groups.append(cur)
        elif cur is not None:
            cur["entries"].append(re.sub(r"^\d+\.\s*", "", s))
    return groups


def main():
    teams = read_rosters()
    apply_trades(teams)
    purses = read_purses()
    prev = load_previous()
    locked = bool(prev and prev.get("locked")) or datetime.now(timezone.utc) >= VALUES_LOCK
    if locked:
        market = frozen_market(prev)
        print("VALUES LOCKED — reusing frozen keeper values (ESPN not queried); "
              "trades/roster edits still apply.")
    else:
        market = fetch_espn()
    owned_ids = set()
    owned_names = set()
    for team in teams:
        team["purse"] = purses.get(team["name"])
        if team["purse"] is None:
            print(f'warning: no {SEASON} auction purse found for {team["name"]}')
    unmatched = []

    for team in teams:
        for p in team["players"]:
            entry, espn_pos = lookup(market, p["name"], p["pos"])
            n = norm_name(p["name"])
            owned_names.add(ALIASES.get(n, n))
            if entry is None:
                unmatched.append(f'{p["name"]} ({p["pos"]}, {team["name"]})')
                mval = float(p["sheetMarket"] or 1)
                p["img"] = player_img(p["name"], p["pos"], None)
                p["nfl"] = p.get("nfl")  # keep any prior value
            else:
                owned_ids.add(entry["id"])
                mval = adp_round(entry["aav"])
                if espn_pos != p["pos"]:
                    print(f'note: {p["name"]} listed as {p["pos"]} in sheet, '
                          f'{espn_pos} on ESPN — using ESPN position')
                    p["pos"] = espn_pos
                p["img"] = player_img(p["name"], p["pos"], entry["id"])
                p["nfl"] = entry.get("nfl") if p["pos"] != "D/ST" else None
            p["market"] = fmt_money(mval)

            is_repeat = str(p["acquired"] or "").startswith("Keeper")
            locked_yr2 = str(p["contractCell"] or "").strip().upper() == "CONTRACT"

            # league convention: fractional prices round down
            if p["priceLocked"]:
                price = math.floor(float(p["sheetPrice"]))
                p["status"] = "contract-yr2" if locked_yr2 else "contract-renewal"
                contract = p["contractCell"] if not locked_yr2 else None
                if isinstance(contract, (int, float)):
                    contract = math.floor(float(contract))
            else:
                price = math.floor(keeper_price(p["draftCost"], mval))
                p["status"] = "market"
                contract = price + 5 if is_repeat else None

            p["price"] = fmt_money(price)
            p["nextYear"] = fmt_money(float(contract)) if isinstance(contract, (int, float)) else None
            p["surplus"] = fmt_money(round(mval - price))
            p["draftCost"] = fmt_money(p["draftCost"]) if p["draftCost"] is not None else None
            for k in ("sheetMarket", "sheetPrice", "priceLocked", "contractCell"):
                p.pop(k)

    changelog = update_changelog(prev, teams)
    out = {
        "season": SEASON,
        "priorSeason": SEASON - 1,
        "generated": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "locked": locked,
        "changeLog": changelog,
        "teams": teams,
        "pool": build_pool(market, owned_ids, owned_names),
        "budgets": build_budgets(purses),
        "trades": TRADES,
    }
    (ROOT / "data").mkdir(exist_ok=True)
    (ROOT / "data" / "keepers.json").write_text(json.dumps(out, indent=2))
    js = "window.LEAGUE_DATA = " + json.dumps(out) + ";\n"
    (ROOT / "js").mkdir(exist_ok=True)
    (ROOT / "js" / "data.js").write_text(js)

    # past-season rosters + narrative trade log — static reference, its own file
    # so the daily cloud refresh (which only round-trips data.js) never carries it
    history = read_history()
    history["tradeHistory"] = read_trade_history()
    (ROOT / "js" / "history.js").write_text(
        "window.LEAGUE_HISTORY = " + json.dumps(history) + ";\n")

    n = sum(len(t["players"]) for t in teams)
    hn = sum(len(t["players"]) for yr in history["byYear"].values() for t in yr)
    print(f"{len(teams)} teams, {n} players. ESPN pool: {len(market)} players.")
    print(f"history: {len(history['seasons'])} seasons ({history['seasons']}), {hn} roster rows.")
    if unmatched:
        print("No ESPN match (kept sheet value):")
        for u in unmatched:
            print("  -", u)


if __name__ == "__main__":
    main()
